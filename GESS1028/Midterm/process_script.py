import openpyxl

# 1. 你的核心逻辑函数
def DecideOctalType(n):
    # 确保输入是整数
    n = int(n)
    N_D = n // 512
    R = n - N_D * 512
    N_C = R // 64
    S = R - N_C * 64
    N_B = S // 8
    N_A = S - N_B * 8
    
    if N_A == max(N_A, N_B, N_C, N_D):
        return '1'
    elif N_B == max(N_A, N_B, N_C, N_D):
        return '2'
    elif N_C == max(N_A, N_B, N_C, N_D):
        return '3'
    else:
        return '4'

# 2. 设置文件名
file_name = 'Survey_Data.xlsx'  # <--- 请在这里修改为你的实际Excel文件名
output_name = 'Processed_Survey_Data.xlsx'

try:
    # 加载 Excel 文件
    wb = openpyxl.load_workbook(file_name)
    
    # 获取当前活动的表格（通常是第一个Sheet）
    sheet = wb.active
    
    print("开始处理数据...")

    # 3. 遍历 C2 到 C55 (在代码中 range 的结束需要+1，所以是 56)
    # Excel中：C列是第3列，D列是第4列
    for row in range(2, 56):
        # 获取 C 列 (column=3) 的数据
        c_cell_value = sheet.cell(row=row, column=3).value
        
        # 检查单元格是否为空，防止报错
        if c_cell_value is not None:
            try:
                # 调用你的函数
                result = DecideOctalType(c_cell_value)
                
                # 将结果写入 D 列 (column=4)
                sheet.cell(row=row, column=4).value = result
            except ValueError:
                print(f"第 {row} 行 C列的数据 '{c_cell_value}' 不是有效的数字，已跳过。")
        else:
            print(f"第 {row} 行 C列为空，已跳过。")

    # 4. 保存结果
    wb.save(output_name)
    print(f"处理完成！结果已保存为: {output_name}")

except FileNotFoundError:
    print(f"错误：找不到文件 '{file_name}'，请确认文件路径正确。")
except Exception as e:
    print(f"发生错误: {e}")